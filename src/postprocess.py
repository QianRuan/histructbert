#encoding=utf-8


import argparse
from datetime import datetime
import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import  PatternFill, Font
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from others.logging import init_logger,logger

val_xents_file = 'validation_xent.json'
test_xents_file = 'test_xent.json'
test_rouges_file = 'test_rouges.json'
test_avg_rouges_file = 'test_avg_rouges.json'
 
metrics=['rouge_1_f_score','rouge_2_f_score','rouge_l_f_score',
                 'rouge_1_recall','rouge_2_recall','rouge_l_recall',
                 'rouge_1_precision','rouge_2_precision','rouge_l_precision']

cnndm_reported_baselines = []
cnndm_reported_baselines.append({'model':'MatchSum(RoBERTa-base)', 'rouge_1_f_score':44.41, 'rouge_2_f_score':20.86, 'rouge_l_f_score':40.55})
cnndm_reported_baselines.append({'model':'MatchSum(BERT-base)', 'rouge_1_f_score':44.22, 'rouge_2_f_score':20.62, 'rouge_l_f_score':40.38})
cnndm_reported_baselines.append({'model':'BERTSUMEXT', 'rouge_1_f_score':43.25, 'rouge_2_f_score':20.24, 'rouge_l_f_score':39.63})
cnndm_reported_baselines.append({'model':'BERTSUMEXT w/o interval embeddings', 'rouge_1_f_score':43.20, 'rouge_2_f_score':20.22, 'rouge_l_f_score':39.59})
cnndm_reported_baselines.append({'model':'BERTSUMEXT (large)', 'rouge_1_f_score':43.85, 'rouge_2_f_score':20.34, 'rouge_l_f_score':39.90})
cnndm_reported_baselines.append({'model':'BERTSUM-lead3', 'rouge_1_f_score':40.42, 'rouge_2_f_score':17.62, 'rouge_l_f_score':48.87})
cnndm_reported_baselines.append({'model':'BERTSUM-oracle', 'rouge_1_f_score':52.59, 'rouge_2_f_score':31.24, 'rouge_l_f_score':36.67})

def str2bool(v):
    if v.lower() in ('yes', 'true', 't', 'y', '1'):
        return True
    elif v.lower() in ('no', 'false', 'f', 'n', '0'):
        return False
    else:
        raise argparse.ArgumentTypeError('Boolean value expected.')


def get_test_rouges_result(model,eval_path):
    
    test_rouges_path = eval_path + test_rouges_file
    test_avg_rouges_path = eval_path + test_avg_rouges_file
        
    #test avg. rouges
    test_avg_rouges_result = {'model':None}
    for m in metrics:
        test_avg_rouges_result.update({m:None})
        
    with open(test_avg_rouges_path, 'r') as f:
        test_avg_rouges = json.load(f)
        
    test_avg_rouges_result['model']=model
    for m in metrics:
        test_avg_rouges_result[m]= round(test_avg_rouges[m] * 100, 2)
    
    # test rouges of top 3 steps of the model
    with open(test_rouges_path, 'r') as f:
        test_rouges = json.load(f)
    
    test_rouges_results =[]
    
    if not (model.split('_')[1]=='oracle' or model.split('_')[1].startswith('lead')):
        for r in test_rouges:
            
            test_rouges_result = {'model':None, 'step':None}
            for m in metrics:
                test_rouges_result.update({m:None})
                
            test_rouges_result['model']=model
          
            test_rouges_result['step']=r[0].split('_')[-1].split('.')[0]
            
            for m in metrics:
                test_rouges_result[m]= round(r[1][m] * 100, 2)
                
            test_rouges_results.append(test_rouges_result)
    else:
        test_rouges_result = {'model':None, 'step':None}
        for m in metrics:
            test_rouges_result.update({m:None})
        
        test_rouges_result['model']=model   
        for m in metrics:
            test_rouges_result[m]= round(test_rouges[m] * 100, 2) 
        test_rouges_results.append(test_rouges_result)
      
    return test_avg_rouges_result, test_rouges_results
        
def get_rouges_df(models):
    
    df_cols = ['model'] + metrics
    df_cols1 = ['model','step'] + metrics
       
    avg_rows = []
    step_rows = []
    
    for model in models:
        
        eval_path = args.models_path + model + '/eval/'
    
        if os.path.exists(eval_path+'DONE') and os.path.exists(args.models_path + model+'/DONE'):
            
            test_avg_rouges_result, test_rouges_results = get_test_rouges_result(model, eval_path)
            
            avg_rows.append(test_avg_rouges_result)
            step_rows = step_rows + test_rouges_results
        
        else:
            if not os.path.exists(args.models_path + model+'/DONE'):
                logger.info("---Training of the model is not finished, skip it-------%s"%(model))
            if not os.path.exists(eval_path+'DONE'):
                logger.info("---Evaluation of the model is not finished, skip it-----%s"%(model))
                
                
    df1 = pd.DataFrame(avg_rows, columns = df_cols) 
    df2 = pd.DataFrame(step_rows, columns = df_cols1)                

    return df1, df2

def check_best_models(df):
    
    best_models = {}
    for m in metrics:
        best_models.update({m:None})
    
    
        
    for m in metrics:
        idxmax = df[m].idxmax()
        model = df['model'][idxmax]
        
        if 'step' in df.columns:
            step = df['step'][idxmax]
        else:
            step=None
        
        best_models[m] = (model, step)
    
    
    
    return best_models
        
        
 
def save_eval_results_to_excel(excelfile,sheetname,df):
   
    if not os.path.isfile(excelfile):
        logger.info('The excel file is not existing, creating a new excel file...'+excelfile)       
        wb = Workbook()
        wb.save(excelfile)
        
        
    wb = load_workbook(excelfile)
    if not (sheetname in wb.sheetnames):
        logger.info('The worksheet is not existing, creating a new worksheet...'+sheetname) 
        ws1 = wb.create_sheet(sheetname)
        ws1.title = sheetname
        wb.save(excelfile)
    
            
    book = load_workbook(excelfile)
    idx = wb.sheetnames.index(sheetname)
    ws=book.get_sheet_by_name(sheetname)
    book.remove(ws)
    book.create_sheet(sheetname, idx)
    writer = pd.ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    
    df.to_excel(writer,sheet_name=sheetname, index = False,header= True)
    writer.save()

def color_the_best_metric(excelfile, sheetname, best_models, color, font):
    wb = load_workbook(excelfile)
    ws = wb[sheetname]
   
    # Create a dictionary of column names
    ColNames = {}
    Current  = 0
    for COL in ws.iter_cols(1, ws.max_column):
        ColNames[COL[0].value] = Current
        Current += 1
 
    # Color best metrics 
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for m in metrics:
            
            if 'step' in ColNames.keys():
                if row_cells[ColNames['model']].value == best_models[m][0] and row_cells[ColNames['step']].value == best_models[m][1]:
                    
                    row_cells[ColNames[m]].fill = PatternFill("solid", fgColor=color)
                    row_cells[ColNames[m]].font = Font(b=font)
            else:
                if row_cells[ColNames['model']].value == best_models[m][0]:
                    
                    row_cells[ColNames[m]].fill = PatternFill("solid", fgColor=color)
                    row_cells[ColNames[m]].font = Font(b=font)
                     
    wb.save(excelfile)
            

    
def copy_result_file(source):
    
    #copy the excelfile  
    now = datetime.now()
    dt_string = now.strftime("%Y-%m-%d-%H-%M")
    dir_path = source[:-5]+'_Copy'
    if not os.path.isdir(dir_path):
     os.mkdir(dir_path)
    target = dir_path +'/Copy'+dt_string +'.xlsx'
    wb = load_workbook(source)
    wb.save(target)
    
    return target

def mark_best_models(best_models, df):
    if 'step' not in df.columns:
        for m in metrics:
            print(df.shape)
            for i in range(df.shape(0)):
                if best_models[m][0] == str(df.iloc[[i]]['model']).split('!')[-1]:
                     df.iloc[[i]]['model']='!'+str(df.iloc[[i]]['model'])
    else:
        for m in metrics:
            for i in range(df.shape(0)):
                if best_models[m][0] == str(df.iloc[[i]]['model']).split('!')[-1] and str(best_models[m][1]) == str(df.iloc[[i]]['step']):
                     df.iloc[[i]]['model']='!'+str(df.iloc[[i]]['model'])
      
def generate_eval_results_overview(args):
    logger.info("=================================================")
    logger.info("Generating evaluation results overview...")
    
    models = os.listdir(args.models_path)
    models = [model for model in models if model.startswith(args.dataset+'_')]
    
    baseline_models = [model for model in models if model.split('_')[1]=='bert'  or model.split('_')[1]=='oracle' or model.split('_')[1].startswith('lead')]
    baseline_models.reverse()
    histruct_models = [model for model in models if model.split('_')[1]=='hs']
    
    logger.info("DATASET: %s"%(args.dataset))
    logger.info("There are %i baseline models"%(len(baseline_models)))
    logger.info("There are %i histruct models"%(len(histruct_models)))
    
    
    df1, df2 = get_rouges_df(baseline_models)
    df3, df4 = get_rouges_df(histruct_models)
  
    df_cols = ['model'] + metrics
    df_cols1 = ['model','step'] + metrics
    df5 = pd.DataFrame(cnndm_reported_baselines, columns = df_cols) 
    df6 = pd.DataFrame(cnndm_reported_baselines, columns = df_cols1) 
    
    df7 = pd.DataFrame([{'model':'REPORTED BASELINES------------'}], columns = df_cols) 
    df8 = pd.DataFrame([{'model':'BASELINES------------'}], columns = df_cols) 
    df9 = pd.DataFrame([{'model':'OUR MODELS------------'}], columns = df_cols) 
    df10 = pd.DataFrame([{'model':'REPORTED BASELINES------------'}], columns = df_cols1) 
    df11 = pd.DataFrame([{'model':'BASELINES------------'}], columns = df_cols1) 
    df12 = pd.DataFrame([{'model':'OUR MODELS------------'}], columns = df_cols1) 
    
    avg_dfs = [df7, df5, df8, df1, df9, df3]
    step_dfs = [df10, df6, df11, df2, df12, df4]
    
    avg_df = pd.concat(avg_dfs)  
    step_df = pd.concat(step_dfs) 
    
    
    result_file = args.models_path+args.dataset+'.eval.results.xlsx'
    avg_sheet = 'avg_rouges'
    step_sheet = 'step_rouges'
    save_eval_results_to_excel(result_file, avg_sheet, avg_df)
    save_eval_results_to_excel(result_file, step_sheet, step_df)
    #make a copy
    cp_result_file = copy_result_file(result_file)
  
    hs_avg_best_models = check_best_models(df3)
    hs_step_best_models = check_best_models(df4)
    color_the_best_metric(result_file, avg_sheet, hs_avg_best_models, color="f0e40a", font=True)
    color_the_best_metric(result_file, step_sheet, hs_step_best_models, color="f0e40a", font=True)
    #color_the_best_metric(cp_result_file, avg_sheet, hs_avg_best_models, color="f0e40a", font=True)
    #color_the_best_metric(cp_result_file, step_sheet, hs_step_best_models, color="f0e40a", font=True)
    
    bert_baseline_models = [model for model in models if model.split('_')[1]=='bert']
    df13,df14 = get_rouges_df(bert_baseline_models)
    bert_avg_best_models = check_best_models(df13)
    bert_step_best_models = check_best_models(df14)
    color_the_best_metric(result_file, avg_sheet, bert_avg_best_models, color="DDDDDD", font=True)
    color_the_best_metric(result_file, step_sheet, bert_step_best_models, color="DDDDDD", font=True)
    #color_the_best_metric(cp_result_file, avg_sheet, bert_avg_best_models, color="DDDDDD", font=True)
    #color_the_best_metric(cp_result_file, step_sheet, bert_step_best_models, color="DDDDDD", font=True)
    
    mark_best_models(bert_avg_best_models,avg_df)
    mark_best_models(hs_avg_best_models,avg_df)
    mark_best_models(bert_step_best_models,step_df)
    mark_best_models(hs_step_best_models,step_df)
    logger.info('avg rouges-------------')
    logger.info(avg_df)
    logger.info('step model rouges------')
    logger.info(step_df)

    logger.info("Generate evaluation results overview...DONE")
    
    #return best step models for plotting summary distribution
    return hs_step_best_models, bert_step_best_models
    
def remove_step_models(args):
    logger.info("=================================================")
    logger.info("Removing step models...")
    models = os.listdir(args.models_path)
    models = [model for model in models if model.startswith(args.dataset+'_')]
    histruct_models = [model for model in models if model.split('_')[1]=='hs']
    
    for model in histruct_models:
        
        eval_path = args.models_path + model + '/eval/'
    
        if os.path.exists(eval_path+'DONE') and os.path.exists(args.models_path + model+'/DONE'):
            files = os.listdir(eval_path)
            summ_files = [file for file in files if file.endswith('.gold')]
            steps = [file.split('.')[1].split('_')[1].split('p')[1] for file in summ_files]
            
            files = os.listdir(args.models_path + model)
            step_models = [file for file in files if file.endswith('.pt')]
            removed_step_models = [model for model in step_models if not model.split('.')[0].split('_')[-1] in steps]
            logger.info("remove %i step models from model %s"%(len(removed_step_models),model))
            for m in removed_step_models:
                path = args.models_path + model + '/' + m
                os.remove(path)
                
        else:
            if not os.path.exists(args.models_path + model+'/DONE'):
                logger.info("---Training of the model is not finished, skip it-------%s"%(model))
            if not os.path.exists(eval_path+'DONE'):
                logger.info("---Evaluation of the model is not finished, skip it-----%s"%(model))
    
    logger.info("Remove step models...DONE")
            
            
def plot_val_xent(args):
    logger.info("=================================================")
    logger.info("Plotting validation loss...")
    models = os.listdir(args.models_path)
    models = [model for model in models if model.startswith(args.dataset+'_')]
    histruct_models = [model for model in models if model.split('_')[1]=='hs']
    
    for model in histruct_models:
        
        eval_path = args.models_path + model + '/eval/'
        
        if os.path.exists(eval_path+'DONE') and os.path.exists(args.models_path + model+'/DONE'):
            
            val_xent_path = eval_path + val_xents_file
            with open(val_xent_path, 'r') as f:
                val_xents = json.load(f)
                
                val_xents_dict = {}
                steps = []
                for v in val_xents:
                    xent = round(v[0],2)
                
                    step = int(v[1].split('/')[-1].split('_')[-1].split('.')[0])#
                    
                    steps.append(step)
                    val_xents_dict.update({step:xent})
                steps.sort()
                xents = [val_xents_dict[step] for step in steps]
                
                png_file = eval_path+'val.xents.png'
                
                plt.plot(steps, xents)
                plt.title('val_xents: '+model)
                plt.ylabel('val_xents', fontsize='large')
                plt.xlabel('step', fontsize='large')
                plt.savefig(png_file, bbox_inches='tight')
                plt.close()      
                
           
        else:
            if not os.path.exists(args.models_path + model+'/DONE'):
                logger.info("---Training of the model is not finished, skip it-------%s"%(model))
            if not os.path.exists(eval_path+'DONE'):
                logger.info("---Evaluation of the model is not finished, skip it-----%s"%(model))
    
    logger.info("Plot validation loss...DONE")
    
def get_prob_dic(modelname, step):
    
    path = args.models_path+modelname+'/eval/'
    file = [file for file in os.listdir(path) if file.endswith('step'+str(step)+'.selectedIdx')][0]
    with open(path+file,'r') as f:
        data = json.load(f)
        flat = [item for sublist in data for item in sublist]  
    prob_dic = []
    for i in range(max(flat)+1):
        prob_dic.append(0)
    for i in set(flat):
        prob_dic[i] = round(flat.count(i)/(len(data[0])*len(data)),2)
    return prob_dic
    
    
def get_best_step_model_prob(best_models):
    
    dic = ({i:list(best_models.values()).count(i) for i in list(best_models.values())})
    total = sum(list(dic.values()))
    best = []
    for model in dic.keys():
        if dic[model]/total > 0.45:
            best.append(model)
            logger.info("Best step model: %s, (%i/%i)"%(model,dic[model],total))
    prob_dics = {}        
    if len(best)==0:
        logger.info("There is no model won on more than 45% of the 9 rouge metrics, %s"%(dic))
    else:       
        for model in best:
            name = model[0]
            step = model[1]
            prob_dic = get_prob_dic(name, step)
            prob_dics.update({name+'.step'+str(step):prob_dic})
    return prob_dics
    
       
def plot_summ_distribution(args, hs_step_best_models, bert_step_best_models):
    logger.info("=================================================")
    logger.info("Plotting summary distribution...")
    prob_dics = {}
    
    oracle = args.dataset+'_oracle'
    
    prob_dics.update({oracle+'.step0': get_prob_dic(oracle, 0)})
    
    
    best_bert_prob_dics = get_best_step_model_prob(bert_step_best_models)
    best_hs_prob_dics = get_best_step_model_prob(hs_step_best_models)
    
    prob_dics.update(best_bert_prob_dics)
    prob_dics.update(best_hs_prob_dics)

    png_file = args.models_path+args.dataset+'.summ.dist.png'
    lens = [len(v) for v in prob_dics.values()]
    max_le = max(lens)
    index = [i for i  in range(max_le)]
    for k in list(prob_dics.keys()):
        if len(prob_dics[k])<max_le:
            prob_dics[k] = prob_dics[k] + [0]*(max_le - len(prob_dics[k]))
       
    df = pd.DataFrame(prob_dics, index=index)
    ax = df.plot.bar(rot=0,title='summary distribution, dataset %s'%(args.dataset))
    ax.set_ylim(0,1)
    ax.set_xlabel("sentence position in source text")
    ax.set_ylabel("propotion of selected sentences")
    ax.get_figure().savefig(png_file)
    
    logger.info("Plot summary distribution...DONE")        
        
    
    


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("-generate_eval_results_overview", type=str2bool, nargs='?',const=True,default=True)
    parser.add_argument("-remove_step_models", type=str2bool, nargs='?',const=True,default=True)
    parser.add_argument("-plot_val_xent", type=str2bool, nargs='?',const=True,default=True)
    parser.add_argument("-plot_summ_distribution", type=str2bool, nargs='?',const=True,default=True)

    
    parser.add_argument("-models_path", default='')
    parser.add_argument("-dataset", default='cnndm', type=str, choices=['cnndm', 'arxiv'])


    args = parser.parse_args()
    
    init_logger(args.models_path+args.dataset+'.postpro.log')
    logger.info("#################################################")

    os.environ['KMP_DUPLICATE_LIB_OK']='True'
  
    if (args.generate_eval_results_overview):
        hs_step_best_models, bert_step_best_models = generate_eval_results_overview(args)
        
    if (args.remove_step_models):
        remove_step_models(args)
    
    if (args.plot_val_xent):
        plot_val_xent(args)
        
    if (args.plot_summ_distribution):
        plot_summ_distribution(args, hs_step_best_models, bert_step_best_models)
        
    
        
    
        

        
        
    
    
  
